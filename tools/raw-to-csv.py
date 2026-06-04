"""
raw-to-csv.py  —  Rohdaten -> CSVParser-kompatible CSV

V1-Scope:
- freigegeben für MSCI-EUR-Indexreihen im aktuellen Finanzwesir-Datenlayer
- schreibt bewusst `EUR` als Suffix
- schreibt Werte mit drei Nachkommastellen
- ist Vorab-Konverter/Validator, kein Ersatz für CSVParser.js

Wird durch den Claude-Skill /raw-to-csv gesteuert. Nie direkt aufrufen.
Unterstuetzt: .xls  .xlsx  .csv
Ausgabe-Format: date;index_value  (Semikolon, Komma-Dezimal, EUR-Suffix)
"""

import sys
import os
import argparse
from datetime import datetime


def load_source(path):
    ext = os.path.splitext(path)[1].lower()

    if ext == ".xls":
        try:
            import xlrd
        except ImportError:
            sys.exit("xlrd fehlt. Installation: pip install xlrd")
        wb = xlrd.open_workbook(path)
        ws = wb.sheet_by_index(0)
        return [
            [str(ws.cell(r, c).value).strip() for c in range(ws.ncols)]
            for r in range(ws.nrows)
        ]

    elif ext == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError:
            sys.exit("openpyxl fehlt. Installation: pip install openpyxl")
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        return [
            [str(cell.value or "").strip() for cell in row]
            for row in ws.iter_rows()
        ]

    elif ext == ".csv":
        import csv
        with open(path, encoding="utf-8-sig") as f:
            sample = f.read(2048)
            f.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
            except csv.Error:
                dialect = csv.excel
            return [row for row in csv.reader(f, dialect)]

    else:
        sys.exit(f"Unbekanntes Format '{ext}'. Unterstuetzt: .xls .xlsx .csv")


def run(args):
    if not os.path.exists(args.source):
        sys.exit(f"Quelldatei nicht gefunden: {args.source}")

    all_rows = load_source(args.source)

    # ── Phase 1: Extraktion ───────────────────────────────────────────────────
    data         = []
    parse_errors = []

    for r in range(args.header_row, len(all_rows)):
        row = all_rows[r]

        if len(row) <= max(args.date_col, args.value_col):
            if len(data) >= args.min_rows:
                break
            continue

        raw_date = row[args.date_col]
        raw_val  = row[args.value_col]

        if not raw_date:
            if len(data) >= args.min_rows:
                break
            continue

        try:
            d = datetime.strptime(raw_date, args.date_format)
        except ValueError:
            if len(data) >= args.min_rows:
                break  # Disclaimer oder leere Zeilen am Ende
            parse_errors.append(
                f"  Zeile {r + 1}: Datum nicht lesbar nach {len(data)} Zeilen: '{raw_date}'"
            )
            continue

        try:
            val = float(raw_val.replace(",", "."))
        except (ValueError, AttributeError):
            parse_errors.append(
                f"  Zeile {r + 1}: Wert nicht numerisch: '{raw_val}'"
            )
            continue

        if val <= 0:
            parse_errors.append(f"  Zeile {r + 1}: Wert <= 0: {val}")
            continue

        data.append((d, val))

    if parse_errors:
        print("EXTRAKTION FEHLGESCHLAGEN:")
        for e in parse_errors:
            print(e)
        sys.exit(1)

    # ── Phase 2: Validierung ──────────────────────────────────────────────────
    all_ok = True

    def check(label, ok, detail=""):
        nonlocal all_ok
        if not ok:
            all_ok = False
        status = "[OK]  " if ok else "[FAIL]"
        suffix = f": {detail}" if detail else ""
        print(f"{status} {label}{suffix}")

    # A: Startdatum
    actual_start = data[0][0].strftime("%Y-%m-%d")
    check(
        "Startdatum",
        actual_start == args.expected_start,
        f"erwartet {args.expected_start}, gefunden {actual_start}"
        if actual_start != args.expected_start else actual_start,
    )

    # B: Zeilenanzahl
    check("Zeilenanzahl", len(data) >= args.min_rows, f"{len(data)} Zeilen")

    # C: Keine Duplikate
    seen, dupes = {}, []
    for d, _ in data:
        key = (d.year, d.month)
        if key in seen:
            dupes.append(d.strftime("%Y-%m"))
        seen[key] = True
    check(
        "Keine Duplikate",
        not dupes,
        f"Duplikate: {dupes}" if dupes else f"{len(data)} eindeutig",
    )

    # D: Lueckenlose Monatsfolge
    gaps = []
    for i in range(1, len(data)):
        prev, curr = data[i - 1][0], data[i][0]
        exp_m = prev.month % 12 + 1
        exp_y = prev.year + (1 if prev.month == 12 else 0)
        if curr.month != exp_m or curr.year != exp_y:
            gaps.append(f"{prev.strftime('%Y-%m')} -> {curr.strftime('%Y-%m')}")
    check(
        "Lueckenlose Monatsfolge",
        not gaps,
        f"Luecken: {gaps}" if gaps else f"alle {len(data)} Monate konsekutiv",
    )

    # E: Alle Werte > 0
    neg = [d[0].strftime("%Y-%m") for d in data if d[1] <= 0]
    check("Alle Werte > 0", not neg, f"{neg}" if neg else "OK")

    if not all_ok:
        print("\nCSV NICHT GESCHRIEBEN — Validierung fehlgeschlagen.")
        sys.exit(1)

    # ── Phase 3: Schreiben ────────────────────────────────────────────────────
    os.makedirs(os.path.dirname(os.path.abspath(args.output)), exist_ok=True)
    with open(args.output, "w", encoding="utf-8") as f:
        f.write("date;index_value\n")
        for d, val in data:
            val_str = f"{val:.3f}".replace(".", ",") + " EUR"
            f.write(f"{d.strftime('%Y-%m-%d')};{val_str}\n")

    print(f"\nCSV geschrieben: {len(data)} Zeilen")
    print(f"Zeitraum:        {data[0][0].strftime('%Y-%m-%d')} bis {data[-1][0].strftime('%Y-%m-%d')}")
    print(f"Ausgabe:         {args.output}")


if __name__ == "__main__":
    p = argparse.ArgumentParser(description="Rohdaten -> CSVParser-kompatible CSV")
    p.add_argument("--source",          required=True,  help="Rohdatei (.xls/.xlsx/.csv)")
    p.add_argument("--output",          required=True,  help="Ziel-CSV-Pfad")
    p.add_argument("--header-row",      type=int, required=True, help="Erste Datenzeile (0-indiziert)")
    p.add_argument("--date-col",        type=int, required=True, help="Spalte Datum (0-indiziert)")
    p.add_argument("--date-format",     required=True,  help="Python-Datumsformat z.B. '%%b %%d, %%Y'")
    p.add_argument("--value-col",       type=int, required=True, help="Spalte Wert (0-indiziert)")
    p.add_argument("--expected-start",  required=True,  help="Erwartetes Startdatum YYYY-MM-DD")
    p.add_argument("--min-rows",        type=int, required=True, help="Mindestanzahl Datenpunkte")
    run(p.parse_args())
